import random
from django.core.files.base import ContentFile 
from django.db.utils import IntegrityError
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import UploadedFile
from django.core.files.storage import default_storage
from openpyxl_image_loader import SheetImageLoader
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from decimal import Decimal, getcontext, ROUND_HALF_UP
from openpyxl.reader.excel import load_workbook
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import models 
from django.db.models.query import QuerySet
import io
import os 
from loguru import logger
from tempfile import NamedTemporaryFile
from .models import Category, Product, Brand, ImportProductsStatus
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from .exceptions import ProductImportError, EndOfTable
from .documents import ProductDocument
from Aleucos import settings
from configs.models import Config
from PIL import Image, ImageDraw, ImageFont
from PIL import Image as PILImage, ImageDraw, ImageFont
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from io import BytesIO
from .models import WatermarkConfig
from dataclasses import dataclass
from datetime import datetime


@dataclass
class WatermarkConfigLocal: 
    font_size: int 
    text: str 
    position: str
    opacity: int


class CatalogImporter:
    @staticmethod
    def import_catalog_from_xlsx(xlsx_file: UploadedFile) -> int:
        workbook = load_workbook(filename=xlsx_file, data_only=True)
        worksheet = workbook.worksheets[1]
        image_loader = SheetImageLoader(worksheet)
        watermark_conf_admin = WatermarkConfig.get_instance()
        watermark_conf = WatermarkConfigLocal(
            font_size=watermark_conf_admin.font_size, 
            text=watermark_conf_admin.text, 
            position=watermark_conf_admin.position,
            opacity=watermark_conf_admin.opacity,
        )

        products_data = []

        for index, row in enumerate(worksheet.iter_rows(min_row=4, values_only=True), 4):
            try:
                product_data = CatalogImporter.process_row(index, row, image_loader, watermark_conf)
                if product_data:
                    products_data.append(product_data)

            except EndOfTable:
                break
            except ProductImportError as e:
                log_text = f'Ошибка при импорте каталога в строке №{index}: {str(e)}. Импорт был прерван'
                logger.error(log_text)
                ImportProductsStatusService.error(log_text)
                return
            
        try:
            CatalogImporter.check_duplicates(products_data)
        except ProductImportError as e:
            log_text = f'Ошибка при импорте каталога: {str(e)}. Импорт был прерван'
            logger.error(log_text)
            ImportProductsStatusService.error(log_text)
            return


        for product_data in products_data:
            CatalogImporter.save_product_data(product_data)
        log_text = f'Каталог был успешно импортирован!'

        logger.success(log_text)
        ImportProductsStatusService.success(log_text)

    @staticmethod 
    def check_duplicates(products_data: list[dict]) -> None: 
        print('Проверяем дубликаты')
        barcodes = [product['barcode'] for product in products_data]
        duplicate_barcodes = list(set([barcode for barcode in barcodes if barcodes.count(barcode) > 1]))
        print(duplicate_barcodes)

        if duplicate_barcodes:
            raise ProductImportError(f'В таблице обнаружены дубликаты штрихкодов: {", ".join(duplicate_barcodes)}')
        
    @staticmethod 
    def save_product_data(product_data: dict) -> None: 
        brand, _ = Brand.objects.get_or_create(title=str(product_data['brand_title'])) 

        product = Product.objects.filter(barcode=product_data['barcode']).first()
        if product: 
            product.remains += product_data['remains']
            product.will_arrive_at = product_data['arriving_date']
            logger.info(f'У товара "{product_data["barcode"]}" обновлён остаток ({product.remains} шт.)')
            product.save()
        else: 
            photo = product_data['photo']

            if photo is None:
                photo = settings.DEFAULT_IMAGE_PATH
            else:
                CatalogImporter.delete_image_if_exists(photo.name)

            category, created = Category.objects.get_or_create(title=product_data['category'])
            if created: 
                logger.info(f'Была добавлена новая категория: {category.title}')

            print(category)

            product = Product(
                barcode=product_data['barcode'],
                brand=brand,
                title=product_data['title'],
                description=product_data['description'],
                photo=product_data['photo'],
                volume=product_data['volume'],
                weight=product_data['weight'],
                notes=product_data['notes'],
                price_before_200k=product_data['price_before_200k'],
                price_after_200k=product_data['price_after_200k'],
                price_after_500k=product_data['price_after_500k'],
                is_in_stock=product_data['is_in_stock'],
                category=category,
                remains=product_data['remains'], 
                will_arrive_at=product_data['arriving_date']
            )
            product.save()
            logger.info(f'Товар "{product_data["title"]}" сохранён в базу данных')

    @staticmethod
    def process_row(index: int, row: tuple, image_loader: SheetImageLoader, watermark_conf: WatermarkConfigLocal) -> None:
        barcode = row[0]
        brand_title = row[1]
        title = row[2]
        description = str(row[3])
        photo = row[4]
        volume = row[5]
        weight = row[6]
        notes = row[7]
        price_before_200k = row[9]
        price_after_200k = row[10]
        price_after_500k = row[11]
        remains = row[16]
        category = row[17]
        arriving_date = None

        if brand_title is None and title is None and barcode is None:
            raise EndOfTable()

        CatalogImporter.validate_product_data(
            barcode,
            title, 
            price_before_200k, 
            price_after_200k, 
            price_after_500k, 
            remains,
            category,
        )

        remains = 0 if remains is None else int(remains)
        is_in_stock = bool(remains)

        if not is_in_stock: 
            arriving_date = row[18] 
            if arriving_date: 
                if type(arriving_date) == str:         
                    try:
                        arriving_date = datetime.strptime(arriving_date, "%d.%m.%Y")
                        print("Дата успешно распаршена:", arriving_date)
                    except ValueError as e:
                        raise ProductImportError(f'Ошибка парсинга даты: {arriving_date}. {str(e)}')
                elif type(arriving_date) != datetime: 
                    arriving_date = None
        
        photo = CatalogImporter.get_image_or_none(barcode, index, image_loader, watermark_conf)

        if weight is not None:
            weight = CatalogImporter.convert_str_to_decimal(str(weight))
        price_before_200k = CatalogImporter.convert_str_to_decimal(str(price_before_200k))
        price_after_200k = CatalogImporter.convert_str_to_decimal(str(price_after_200k))
        price_after_500k = CatalogImporter.convert_str_to_decimal(str(price_after_500k))

        product_data = {
            'barcode': barcode,
            'brand_title': brand_title,
            'title': title,
            'description': description,
            'photo': photo,
            'volume': volume,
            'weight': weight,
            'notes': notes,
            'price_before_200k': price_before_200k,
            'price_after_200k': price_after_200k,
            'price_after_500k': price_after_500k,
            'is_in_stock': is_in_stock,
            'category': category,
            'remains': remains, 
            'arriving_date': arriving_date,
        }

        return product_data

    @staticmethod
    def get_image_or_none(barcode: str, row_index: int, image_loader: SheetImageLoader, watermark_conf: WatermarkConfigLocal) -> ContentFile | None:
        try:
            image = image_loader.get(f'E{row_index}')
            image_stream = io.BytesIO()
            image.save(image_stream, format='PNG')
            image_stream.seek(0)

            watermarked_image = add_watermark(
                image_stream, 
                barcode, 
                text=watermark_conf.text, 
                font_size=watermark_conf.font_size, 
                position=watermark_conf.position, 
                opacity=watermark_conf.opacity
            )

            if watermarked_image:
                logger.info(f"Водяной знак успешно добавлен на изображение для товара с баркодом {barcode}.")
            else:
                logger.warning(f"Не удалось добавить водяной знак для товара с баркодом {barcode}.")

            return watermarked_image
        except Exception as e:
            logger.error(str(e))
            return None

    @staticmethod
    def delete_image_if_exists(image_name: str) -> bool:
        image_path = os.path.join('products', image_name)
        if default_storage.exists(image_path):
            default_storage.delete(image_path)

    @staticmethod
    def convert_str_to_decimal(value: str) -> Decimal:
        getcontext().clamp = 1
        value = Decimal(float(str(value).replace(',', '.')))
        return value.quantize(Decimal('0.000001'), rounding=ROUND_HALF_UP)

    @staticmethod
    def validate_product_data(
        barcode: str | float | None,
        title: str | None,
        price_before_200k: float | None,
        price_after_200k: float | None,
        price_after_500k: float | None, 
        remains: int | str | None, 
        category: str | None
    ) -> None:
        if title is None:
            raise ProductImportError(f'У товара со штрихкодом {barcode} отсутствует название')
        elif barcode is None or str(barcode) == '0':
            raise ProductImportError(f'У товара {title} отсутствует штрихкод')
        elif any(price in (None, 0) for price in (price_before_200k, price_after_200k, price_after_500k)):
            raise ProductImportError(f'У товара {title} со штрихкодом {barcode} отсутствует цена')
        elif not str(barcode).strip().isnumeric():
            raise ProductImportError(f'У товара неверный штрихкод: {barcode}')
        elif str(category) == "" or category is None:
            raise ProductImportError(f'У товара со штрихкодом {barcode} не указана категория')
        if remains: 
            if not str(remains).strip().isnumeric(): 
                raise ProductImportError(f'У товара неверный остаток на складе: {remains}')

    @staticmethod
    def truncate_products_and_brands() -> None:
        Product.objects.all().delete()
        Brand.objects.all().delete()


def add_watermark(image_path, barcode: str, text="©Aleucos", font_size=30, position="bottom_left", opacity=255):
    image = PILImage.open(image_path).convert("RGBA")
    txt_layer = PILImage.new("RGBA", image.size, (255, 255, 255, 0))
    draw = ImageDraw.Draw(txt_layer)
    font = ImageFont.load_default()
    text_width, text_height = draw.textbbox((0, 0), text)[2:]
    positions = {
        "top_left": (10, 10),
        "top_right": (image.width - text_width - 10, 10),
        "bottom_left": (10, image.height - text_height - 10),
        "bottom_right": (image.width - text_width - 10, image.height - text_height - 10),
        "center": ((image.width - text_width) // 2, (image.height - text_height) // 2)
    }
    text_position = positions.get(position, positions["bottom_left"])
    shadow_offset = (1, 1)
    draw.text((text_position[0] + shadow_offset[0], text_position[1] + shadow_offset[1]), text, font=font, fill=(0, 0, 0, opacity // 2))
    draw.text(text_position, text, font=font, fill=(255, 255, 255, opacity))
    watermarked_image = PILImage.alpha_composite(image, txt_layer)
    
    output_stream = BytesIO()
    watermarked_image.convert("RGB").save(output_stream, "PNG")
    output_stream.seek(0)

    return ContentFile(output_stream.read(), f"{barcode}.png")


class CatalogExporter:
    @staticmethod
    def export_catalog_to_xlsx() -> str:
        catalog_template_path = os.path.join(settings.MEDIA_ROOT, 'catalog', Config.get_instance().export_catalog_template_filename)
        exported_catalog_path = os.path.join(settings.MEDIA_ROOT, 'catalog', Config.get_instance().export_catalog_filename)

        try:
            workbook = load_workbook(filename=catalog_template_path, data_only=True)
        except PermissionError:
            logger.error('Ошибка при экспорте каталога: файл занят другим процессом')
            return ""
        
        worksheet = workbook['Актуальное наличие на складе']

        products = Product.objects.all()
        curr_row_index = 4

        for product in products:
            worksheet.merge_cells(f'T{curr_row_index}:W{curr_row_index}')
            
            worksheet[f'A{curr_row_index}'] = str(product.barcode)
            worksheet[f'B{curr_row_index}'] = product.brand.title
            worksheet[f'C{curr_row_index}'] = product.title
            worksheet[f'D{curr_row_index}'] = product.description
            worksheet[f'F{curr_row_index}'] = product.volume
            worksheet[f'G{curr_row_index}'] = product.weight
            worksheet[f'H{curr_row_index}'] = product.notes
            worksheet[f'J{curr_row_index}'] = product.price_before_200k
            worksheet[f'K{curr_row_index}'] = product.price_after_200k
            worksheet[f'L{curr_row_index}'] = product.price_after_500k
            worksheet[f'Q{curr_row_index}'] = product.remains
            worksheet[f'R{curr_row_index}'] = product.category.title
            worksheet[f'S{curr_row_index}'] = product.will_arrive_at

            image_path = os.path.join(settings.MEDIA_ROOT, product.photo.name)
            try:
                img = Image(image_path)

                img.width = 102
                img.height = 100

                worksheet.add_image(img, f"E{curr_row_index}")

                worksheet.column_dimensions['E'].width = 15 
                worksheet.row_dimensions[curr_row_index].height = 80
            except FileNotFoundError:
                worksheet[f'E{curr_row_index}'] = "Файл не найден"

            curr_row_index += 1

        worksheet['J2'] = f"=SUMPRODUCT(J4:J{curr_row_index - 1}, M4:M{curr_row_index - 1})"
        worksheet['K2'] = f"=SUMPRODUCT(K4:K{curr_row_index - 1}, M4:M{curr_row_index - 1})"
        worksheet['L2'] = f"=SUMPRODUCT(L4:L{curr_row_index - 1}, M4:M{curr_row_index - 1})"

        temp_file_path = None
        try:
            with NamedTemporaryFile(delete=False, suffix='.xlsx', dir=os.path.join(settings.MEDIA_ROOT, 'tmp')) as temp_file:
                workbook.save(temp_file.name)
                temp_file_path = temp_file.name
            os.replace(temp_file_path, exported_catalog_path)
        except Exception as e:
            logger.error(f'Ошибка при сохранении файла: {e}')
        finally:
            if temp_file_path and os.path.exists(temp_file_path):
                os.remove(temp_file_path)

        return exported_catalog_path



class ElasticSearchService: 
    @staticmethod
    def truncate_products_index() -> None:
        ProductDocument().search().query('match_all').delete() 

    @staticmethod 
    def add_all_products_to_index() -> None: 
        ProductDocument().update(Product.objects.all())


class ImportProductsStatusService: 
    @staticmethod 
    def info(text: str) -> None: 
        ImportProductsStatus.objects.create(text=text, status_type=ImportProductsStatus.Type.INFO).save()

    @staticmethod 
    def process(text: str) -> None: 
        ImportProductsStatus.objects.create(text=text, status_type=ImportProductsStatus.Type.PROCESS).save()

    @staticmethod 
    def error(text: str) -> None: 
        ImportProductsStatus.objects.create(text=text, status_type=ImportProductsStatus.Type.ERROR).save()

    @staticmethod 
    def success(text: str) -> None: 
        ImportProductsStatus.objects.create(text=text, status_type=ImportProductsStatus.Type.SUCCESS).save()

    @staticmethod 
    def get_all_statuses() -> QuerySet[ImportProductsStatus]: 
        return ImportProductsStatus.objects.all().order_by('time')
    
    @staticmethod 
    def delete_all() -> None: 
        ImportProductsStatus.objects.all().delete()


def get_max_product_price() -> Decimal: 
    max_price = Product.objects.aggregate(models.Max('price_before_200k'))['price_before_200k__max']
    return max_price


def get_max_product_weight() -> Decimal: 
    max_weight = Product.objects.aggregate(models.Max('weight'))['weight__max']
    return max_weight

 
def get_paginated_collection(request, collection: QuerySet, count_per_page: int = 10): 
    paginator = Paginator(collection, count_per_page)
    page_number = request.GET.get('page', 1)
    try:
        collection = paginator.page(page_number)
    except PageNotAnInteger:
        collection = paginator.page(1)
    except EmptyPage:
        collection = paginator.page(paginator.num_pages) 
    return collection


def get_all_model_objects(model: models.Model) -> QuerySet: 
    return model.objects.all()
