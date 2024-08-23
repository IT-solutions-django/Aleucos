from django.contrib import admin
from django.urls import path
from django.shortcuts import redirect, render
from django.db.utils import IntegrityError
from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile 
from openpyxl_image_loader import SheetImageLoader
from decimal import Decimal, getcontext, ROUND_HALF_UP
from openpyxl.reader.excel import load_workbook
import io

from .forms import XlsxImportForm
from .models import Brand, Product 


@admin.register(Brand)
class BrandAdmin(admin.ModelAdmin): 
    list_display = ['pk', 'title']
    list_filter = ['title']


@admin.register(Product) 
class ProductAdmin(admin.ModelAdmin): 
    list_display = ['barcode', 'brand', 'title', 'volume', 'weight', 'photo',
                    'price_before_200k', 'price_after_200k', 'price_after_500k'] 
    list_filter = ['brand', 'volume', 'weight', 
                    'price_before_200k', 'price_after_200k', 'price_after_500k'] 
    search_fields = ['brand__title', 'title']

    change_list_template = 'products/product_change_list.html'


    def get_urls(self):
        urls = super().get_urls()

        my_urls = [
            path('import-products-from-xlsx/', self.import_products_from_xlsx),
        ]
        return my_urls + urls

    def import_products_from_xlsx(self, request):
        context = admin.site.each_context(request)
        if request.method == 'POST':

            xlsx_file = request.FILES['xlsx_file']
            workbook = load_workbook(filename=xlsx_file, data_only=True)
            worksheet = workbook.worksheets[1]
            image_loader = SheetImageLoader(worksheet)

            updated_products_count = 0
            for i, row in enumerate(worksheet.iter_rows(min_row=4, values_only=True), 4):
                barcode = row[0]
                brand_title = row[1]
                title = row[2]
                description = row[3]
                photo = row[4]
                volume = row[5]
                weight = row[6]
                notes = row[7]
                price_before_200k = row[9]
                price_after_200k = row[10]
                price_after_500k = row[11]
                is_in_stock = True if str(row[12]) == '0' else False

                if not brand_title and not title and not barcode:
                    break
                if not self.validate_row(request, title, barcode, price_before_200k, price_after_200k, price_after_500k): 
                    continue

                try:
                    image = image_loader.get(f'E{i}')
                    image_stream = io.BytesIO()
                    image.save(image_stream, format='PNG')
                    image_stream.seek(0)
                    photo = ContentFile(image_stream.read(), f'{barcode}.png')
                except (KeyError, ValueError):
                    photo = None  

                getcontext().clamp = 1
                if weight is not None:
                    weight_decimal = Decimal(str(weight).replace(',', '.'))
                price_before_200k = Decimal(price_before_200k).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                price_after_200k = Decimal(price_after_200k).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                price_after_500k = Decimal(price_after_500k).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                brand, _ = Brand.objects.get_or_create(title=brand_title)
                
                try:
                    Product.objects.update_or_create(
                        barcode=barcode, 
                        defaults={
                            'brand': brand,
                            'title': title,
                            'description': description,
                            'photo': photo,
                            'volume': volume,
                            'weight': weight_decimal,
                            'notes': notes,
                            'price_before_200k': price_before_200k,
                            'price_after_200k': price_after_200k,
                            'price_after_500k': price_after_500k,
                            'is_in_stock': is_in_stock
                        }
                    )
                    updated_products_count += 1

                except IntegrityError as e: 
                    self.message_user(request, f'Ошибка импорта: {e}', level='error')
                except ValidationError as e: 
                    self.message_user(request, f'Ошибка валидации: {e}', level='error')
                except TypeError as e: 
                    self.message_user(request, f'Неверный тип данных: {e}', level='error')

            self.message_user(request, f'Добавлено или обновлено {updated_products_count} товаров')
            return redirect('admin:products_product_changelist')

        context['form'] = XlsxImportForm()
        return render(request, 'products/add_products_form.html', context=context)
    
    def validate_row(self, request, title, barcode, price_before_200k, price_after_200k, price_after_500k) -> bool: 
        if not title:
            self.message_user(request, f'У товара со штрихкодом {barcode} отсутствует название', level='error')
            return False
        elif price_before_200k is None or price_after_200k is None or price_after_500k is None: 
            self.message_user(request, f'У товара {title} со штрихкодом {barcode} отсутствует цена', level='error')
            return False
        elif not barcode or str(barcode) == '0': 
            self.message_user(request, f'У товара {title} отсутствует штрихкод', level='error')
            return False
        elif not str(barcode).strip().isnumeric(): 
            self.message_user(request, f'У товара неверный штрихкод: {barcode}', level='error')
            return False

        return True