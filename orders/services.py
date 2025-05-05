from django.db.utils import IntegrityError
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import UploadedFile
from decimal import Decimal, getcontext, ROUND_HALF_UP
from openpyxl.reader.excel import load_workbook
from loguru import logger
from django.db.models.query import QuerySet
from .exceptions import OrderImportError, EndOfTable
from .models import OrderStatus, Order, OrderItem, ImportOrderStatus
from products.models import Product
from Aleucos import settings
from users.models import User
from configs.models import Config
from orders.models import PaymentMethod, DeliveryTerm
from users.models import City
from openpyxl.reader.excel import load_workbook 
from products.models import Product
from tempfile import NamedTemporaryFile
from openpyxl.drawing.image import Image
from django.core.files.base import ContentFile
from openpyxl.utils.units import pixels_to_EMU
from io import BytesIO 
import os
from Aleucos.elastic_log_handler import log_product_sale
from openpyxl.styles import Border, Side
from copy import copy
import openpyxl

from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.utils import column_index_from_string



class OrderImporter:
    @staticmethod
    def import_order_from_xlsx(
        xlsx_file: UploadedFile, 
        manager_email: str,
        payment_method_id: int, 
        delivery_terms_id: int, 
        city: str,
        comment: str, 
        user_id: int
    ) -> None:
        workbook = load_workbook(filename=xlsx_file, data_only=True)
        items_worksheet = workbook.worksheets[1]

        user = User.objects.get(pk=user_id)
        
        try: 
            manager = User.objects.get(email=manager_email)
        except User.DoesNotExist: 
            log_text = f'Ошибка! Заказ не был создан: менеджера с email {manager_email} не существует'
            ImportOrderStatusService.error(log_text, manager_email)
            logger.error(log_text)
            return 

        order_data = {
            'user': user,
            'status': OrderStatus.objects.get(title=Config.get_instance().order_status_first),
            'manager': manager,
            'items': [],
            'total_price': 0
        }

        user_discount = user.discount / 100 if user.discount else 0
        final_price_coefficient = (1 - user_discount)

        for index, row in enumerate(items_worksheet.iter_rows(min_row=4, values_only=True), 4):
            try:
                order_item_data = OrderImporter.process_order_row(index, row)
                if order_item_data:
                    order_data['items'].append(order_item_data)
            except EndOfTable:
                break
            except OrderImportError as e:
                log_text = f'Ошибка! Заказ не был создан: {str(e)}'
                logger.error(log_text)
                ImportOrderStatusService.error(log_text, manager_email)
                return 
            
        if len(order_data['items']) == 0: 
            log_text = f'Ошибка! Заказ не был создан: в заказе нет ни одной позиции'
            ImportOrderStatusService.error(log_text, manager_email)
            logger.error(log_text)
            return 

        OrderImporter.calculate_total_price(order_data)

        order = Order.objects.create(
            user=order_data['user'],
            status=order_data['status'],
            manager=order_data['manager'],
            total_price=float(order_data['total_price']) * final_price_coefficient, 
            payment_method=PaymentMethod.objects.get(pk=payment_method_id), 
            delivery_terms=DeliveryTerm.objects.get(pk=delivery_terms_id), 
            city=city,
            comment=comment
        )


        articles = []
        for item_data in order_data['items']: 
            articles.append(item_data['article'])
        products = Product.objects.filter(article__in=articles)
        products_dict = {p.article: p for p in products}  


        order_items = []
        for item_data in order_data['items']:
            order_items.append(
                OrderItem(
                    article=item_data['article'],
                    order=order,
                    product_name=item_data['product_name'], 
                    brand_name=item_data['brand_name'], 
                    quantity=item_data['quantity'], 
                    unit_price=float(item_data['unit_price']) * final_price_coefficient, 
                    total_price=float(item_data['total_price']) * final_price_coefficient
                )
            )
            product = products_dict[item_data['article']]
            product.remains -= item_data['quantity']

            log_product_sale(
                product=product, 
                quantity=item_data['quantity'], 
                manager_name=manager.get_fullname()
            )

            product.save()
            
            log_text = f'В заказ добавлен товар {item_data["product_name"]} ({item_data["quantity"]} шт)'
            logger.info(log_text)
            ImportOrderStatusService.process(log_text, manager_email)

        OrderItem.objects.bulk_create(order_items)
        order.save()
        order.create_pdf_bill()

        log_text = f'Заказ №{order.number} для клиента {order.user.email} был успешно создан!' if order.user else f'Заказ №{order.number} был успешно создан!'
        logger.info(log_text)
        ImportOrderStatusService.success(log_text, manager_email)

    @staticmethod
    def process_order_row(index: int, row: tuple) -> OrderItem | None:
        article = row[0]
        barcode = row[1]
        brand_title = row[2]
        title = row[3]
        quantity = int(row[16]) if row[16] is not None else 0 

        if quantity == 0: 
            return 
        if brand_title is None and title is None and article is None:
            raise EndOfTable()
        OrderImporter.validate_product_data(article, title)

        try:
            product = Product.objects.get(article=article)
        except Product.DoesNotExist: 
            raise OrderImportError(f'Товара со артикулом {article} не существует')
        if product.remains - quantity < 0: 
            raise OrderImportError(f'Товара со артикулом {article}  недостаточно на складе ({product.remains} шт. есть, запрашивается {quantity} шт.)')

        price_before_200k = product.price_before_200k

        unit_price = price_before_200k  
        total_price = quantity * unit_price

        return {
            'article': product.article,
            'product_name': product.title,
            'brand_name': product.brand.title,
            'quantity': quantity,
            'unit_price': unit_price,
            'total_price': total_price
        }
    
    @staticmethod
    def calculate_total_price(order_data: dict) -> None:
        total_price = sum(item['total_price'] for item in order_data['items'])

        if total_price >= 200000:
            for item in order_data['items']:
                item['unit_price'] = Product.objects.get(article=item['article']).price_after_200k
                item['total_price'] = item['quantity'] * item['unit_price']
            total_price = sum(item['total_price'] for item in order_data['items'])

        if total_price >= 500000:
            for item in order_data['items']:
                item['unit_price'] = Product.objects.get(article=item['article']).price_after_500k
                item['total_price'] = item['quantity'] * item['unit_price']
            total_price = sum(item['total_price'] for item in order_data['items'])

        order_data['total_price'] = total_price

    @staticmethod
    def convert_str_to_decimal(value: str) -> Decimal:
        getcontext().clamp = 1
        value = Decimal(float(str(value).replace(',', '.')))
        return value.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    @staticmethod
    def validate_product_data(article: str | float | None, title: str | None) -> None:
        if article is None: 
            raise OrderImportError(f'У товара отсутствует артикул')
        if title is None:
            raise OrderImportError(f'У товара со артикулом {article} отсутствует название')       


class ImportOrderStatusService: 
    @staticmethod 
    def info(text: str, manager_email: str) -> None: 
        ImportOrderStatus.objects.create(
            text=text, 
            status_type=ImportOrderStatus.Type.INFO, 
            manager=User.objects.get(email=manager_email)
        ).save()

    @staticmethod 
    def process(text: str, manager_email: str) -> None: 
        ImportOrderStatus.objects.create(
            text=text, 
            status_type=ImportOrderStatus.Type.PROCESS,
            manager=User.objects.get(email=manager_email)
        ).save()

    @staticmethod 
    def error(text: str, manager_email: str) -> None: 
        ImportOrderStatus.objects.create(
            text=text, 
            status_type=ImportOrderStatus.Type.ERROR, 
            manager=User.objects.get(email=manager_email)
        ).save()

    @staticmethod 
    def success(text: str, manager_email: str) -> None: 
        ImportOrderStatus.objects.create(
            text=text, 
            status_type=ImportOrderStatus.Type.SUCCESS, 
            manager=User.objects.get(email=manager_email)
        ).save()

    @staticmethod 
    def delete_all() -> None: 
        ImportOrderStatus.objects.all().delete()

    @staticmethod 
    def get_all_statuses(manager: User) -> QuerySet | None: 
        return ImportOrderStatus.objects.all().filter(manager=manager)


class OrderExcelGenerator:
    @staticmethod
    def export_order_to_xlsx(order: Order) -> str:
        print(f'Количество товаров: {order.items.all().count()}')
        catalog_template_path = os.path.join(settings.MEDIA_ROOT, 'catalog', 'order_template.xlsx')

        try:
            workbook = load_workbook(filename=catalog_template_path, data_only=True)
        except PermissionError:
            logger.error('Ошибка при экспорте каталога: файл занят другим процессом')
            return ""
        
        worksheet = workbook['Актуальное наличие на складе']

        items = order.items.all()

        curr_row_index = 4

        for item in items:
            worksheet.merge_cells(f'U{curr_row_index}:X{curr_row_index}')

            item: OrderItem

            product = Product.objects.get(article=item.article)
            
            worksheet[f'A{curr_row_index}'] = str(product.article)
            worksheet[f'B{curr_row_index}'] = str(product.barcode)
            worksheet[f'C{curr_row_index}'] = product.brand.title
            worksheet[f'D{curr_row_index}'] = product.title
            worksheet[f'E{curr_row_index}'] = product.title_russian if (product.title_russian is not None and product.title_russian != 'None') else ''
            worksheet[f'G{curr_row_index}'] = product.volume
            worksheet[f'H{curr_row_index}'] = product.weight
            worksheet[f'I{curr_row_index}'] = product.notes

            worksheet[f'K{curr_row_index}'] = product.remains
            worksheet[f'L{curr_row_index}'] = ','.join(category.title for category in product.categories.all())
            worksheet[f'M{curr_row_index}'] = product.will_arrive_at

            worksheet[f'N{curr_row_index}'] = product.price_before_200k
            worksheet[f'O{curr_row_index}'] = product.price_after_200k
            worksheet[f'P{curr_row_index}'] = product.price_after_500k
            
            worksheet[f'Q{curr_row_index}'] = item.quantity

            if product.photo:
                image_path = os.path.join(settings.MEDIA_ROOT, product.photo.name)
                try:
                    img = Image(image_path)
                    proto_image = copy(img)

                    # Максимальные размеры
                    MAX_WIDTH = 160
                    MAX_HEIGHT = 100

                    # Получаем текущие размеры
                    current_width = proto_image.width
                    current_height = proto_image.height

                    # Проверяем, нужно ли масштабировать
                    if current_width > MAX_WIDTH or current_height > MAX_HEIGHT:
                        # Определяем, по какому измерению нужно масштабировать
                        width_ratio = MAX_WIDTH / current_width
                        height_ratio = MAX_HEIGHT / current_height
                        
                        # Выбираем меньший коэффициент масштабирования
                        scale_ratio = min(width_ratio, height_ratio)
                        
                        # Применяем масштабирование
                        proto_image.width = int(current_width * scale_ratio)
                        proto_image.height = int(current_height * scale_ratio)
                    else:
                        # Если размеры меньше максимальных, оставляем как есть
                        proto_image.width = current_width
                        proto_image.height = current_height

                    col_letter = 'F'

                    # Преобразуем размеры в EMU
                    width_emu = pixels_to_EMU(proto_image.width)
                    height_emu = pixels_to_EMU(proto_image.height)

                    size = XDRPositiveSize2D(width_emu, height_emu)

                    col_index = column_index_from_string(col_letter) - 1

                    # Получаем ширину колонки в символах
                    col_width_chars = worksheet.column_dimensions[col_letter].width or 8.43  # 8.43 - ширина по умолчанию

                    # Преобразуем ширину колонки в пиксели (1 символ ≈ 7.5 пикселей)
                    col_width_px = col_width_chars * 7.5

                    # Центровка: сколько отступа слева
                    left_padding_px = max((col_width_px - proto_image.width) / 2, 0)
                    col_offset_emu = pixels_to_EMU(left_padding_px)

                    # Вычисляем вертикальное центрирование
                    row_height_px = worksheet.row_dimensions[curr_row_index].height or 15  # 15 - стандартная высота строки
                    row_height_px = row_height_px * 1.2  # Конвертируем в пиксели (1.2 - примерный коэффициент)
                    vertical_padding_px = max((row_height_px - proto_image.height) / 2, 0)
                    
                    # Фиксированный отступ сверху
                    TOP_PADDING = 5
                    row_offset_emu = pixels_to_EMU(vertical_padding_px + TOP_PADDING)

                    marker = AnchorMarker(col=col_index, colOff=col_offset_emu, row=curr_row_index - 1, rowOff=row_offset_emu)
                    proto_image.anchor = OneCellAnchor(_from=marker, ext=size)

                    worksheet.add_image(proto_image)

                except FileNotFoundError:
                    worksheet[f'E{curr_row_index}'] = "Файл не найден"

            curr_row_index += 1
        
        thin_border = Border(
            top=Side(style='thin'),
            right=Side(style='thin'),
            bottom=Side(style='thin'),
            left=Side(style='thin')
        )
        for row in worksheet.iter_rows(min_row=4, max_row=curr_row_index-1):
            for cell in row[:17]:
                cell.border = thin_border


        worksheet['N2'] = f"=SUMPRODUCT(N4:N{curr_row_index - 1}, Q4:Q{curr_row_index - 1})"
        worksheet['O2'] = f"=SUMPRODUCT(O4:O{curr_row_index - 1}, Q4:Q{curr_row_index - 1})"
        worksheet['P2'] = f"=SUMPRODUCT(P4:P{curr_row_index - 1}, Q4:Q{curr_row_index - 1})"

        worksheet[f'A1'] = f'Менеджер:'
        worksheet[f'B1'] = f'{order.manager.get_fullname()}'

        worksheet[f'A2'] = f'Клиент:'
        worksheet[f'B2'] = f'{order.user.get_fullname()}'

        output = BytesIO()
        workbook.save(output)
        output.seek(0)  

        filename = f"order_{order.number}.xlsx"

        order.info_excel.save(filename, ContentFile(output.getvalue()))
        order.save()

        output.close()
        return True